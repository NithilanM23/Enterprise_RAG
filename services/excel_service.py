"""
services/excel_service.py
--------------------------
Excel file ingestion, row storage, and lookup for the Local Employee Knowledge Assistant.

Excel is fundamentally different from prose documents — the unit of meaning
is a row, not a paragraph. This service handles Excel separately:

  Ingestion:
    Each row stored as JSONB in excel_rows table.
    Each row also stringified for BM25 keyword search.

  Lookup (Type 1 — row search):
    User asks: "Is order 1042 in the sheet?" / "Find Ravi Kumar's invoice"
    → BM25 search on stringified rows + optional SQL filter for exact IDs
    → Returns matching rows as a table

  Aggregation (Type 2 — computed answers):
    User asks: "How many pending orders?" / "Total sales for March"
    → LLM generates SQL → executed against excel_rows table
    → Returns computed result + LLM natural language summary

  Metadata (Type 3 — structure questions):
    Handled by metadata_service.py — sheet names, row count, column names
    stored at ingestion time.

Schema:
  excel_rows
    id, document_id, sheet_name, row_number, row_data (JSONB), row_text (TEXT)
"""

import json
import logging
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Table creation
# ---------------------------------------------------------------------------

def ensure_excel_table() -> None:
    """Create excel_rows table if it does not exist. Idempotent."""
    from services.database_service import _get_raw_connection

    with _get_raw_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS excel_rows (
                    id          SERIAL PRIMARY KEY,
                    document_id INTEGER  NOT NULL
                                    REFERENCES documents(id) ON DELETE CASCADE,
                    sheet_name  TEXT     NOT NULL,
                    row_number  INTEGER  NOT NULL,
                    row_data    JSONB    NOT NULL,
                    row_text    TEXT     NOT NULL,
                    CONSTRAINT excel_rows_unique UNIQUE (document_id, sheet_name, row_number)
                );
            """)
            # Full-text search index on row_text for fast keyword search
            cur.execute("""
                CREATE INDEX IF NOT EXISTS excel_rows_text_idx
                ON excel_rows USING gin(to_tsvector('english', row_text));
            """)

    logger.debug("excel_rows table and index ensured.")


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def ingest_excel(source, document_id: int, filename: str = "") -> dict:
    """
    Load an Excel file and store all rows in the excel_rows table.

    Args:
        source      : BytesIO object OR filepath string.
                      BytesIO is preferred on Windows to avoid file-lock issues.
        document_id : The document's DB row ID.
        filename    : Original filename (for logging).

    Returns:
        dict with sheet_count, total_rows, sheets (list of sheet summaries)
    """
    import io
    import openpyxl
    import psycopg2.extras
    from services.database_service import _get_connection

    # Safety net — create table if it doesn't exist yet
    # (handles cases where cache_resource skipped initialize_database)
    ensure_excel_table()

    # Always work from BytesIO — avoids Windows file handle leaks
    if isinstance(source, (str, bytes.__class__)) and not isinstance(source, io.IOBase):
        with open(source, "rb") as f:
            data = f.read()
        source = io.BytesIO(data)
    elif hasattr(source, "seek"):
        source.seek(0)

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    total = 0
    sheets_summary = []

    with _get_connection() as conn:
        with conn.cursor() as cur:
            for sheet_name in wb.sheetnames:
                ws   = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    continue

                # First row = headers
                headers  = [str(c).strip() if c is not None else f"col_{i}"
                            for i, c in enumerate(rows[0])]
                data_rows = rows[1:]
                sheet_row_count = 0

                for row_idx, row in enumerate(data_rows, start=1):
                    # Skip completely empty rows
                    if all(v is None for v in row):
                        continue

                    # Build row dict — header: value
                    row_dict = {}
                    for h, v in zip(headers, row):
                        if v is not None:
                            row_dict[h] = v if not hasattr(v, 'isoformat') else str(v)

                    if not row_dict:
                        continue

                    # Stringify for BM25 / full-text search
                    row_text = _stringify_row(row_dict)

                    cur.execute("""
                        INSERT INTO excel_rows
                            (document_id, sheet_name, row_number, row_data, row_text)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (document_id, sheet_name, row_number) DO UPDATE
                            SET row_data = EXCLUDED.row_data,
                                row_text = EXCLUDED.row_text;
                    """, (
                        document_id,
                        sheet_name,
                        row_idx,
                        json.dumps(row_dict),
                        row_text,
                    ))
                    sheet_row_count += 1
                    total += 1

                sheets_summary.append({
                    "sheet_name": sheet_name,
                    "headers":    headers,
                    "row_count":  sheet_row_count,
                })

    try:
        wb.close()
    except Exception:
        pass
    finally:
        import gc
        gc.collect()   # release any lingering Windows handles

    logger.info(
        "Excel ingested: %d sheets, %d rows stored for document_id=%d.",
        len(sheets_summary), total, document_id,
    )
    return {
        "sheet_count": len(sheets_summary),
        "total_rows":  total,
        "sheets":      sheets_summary,
    }


def _stringify_row(row_dict: dict) -> str:
    """
    Convert a row dict to a searchable string.
    Format: "Column1: value1  Column2: value2  ..."
    """
    return "  ".join(f"{k}: {v}" for k, v in row_dict.items())


# ---------------------------------------------------------------------------
# Row lookup (Type 1)
# ---------------------------------------------------------------------------

def search_rows(
    query: str,
    document_ids: list = None,
    top_k: int = 20,
) -> list:
    """
    Search Excel rows using a 3-tier strategy:
      Tier 1: plainto_tsquery  — phrase-based full-text search (lenient, no &-joining)
      Tier 2: ILIKE per term  — partial substring match on row_text
      Tier 3: show all rows   — when query is broad ("show all", "list all", "all rows")

    plainto_tsquery is used instead of to_tsquery because:
      - to_tsquery with & requires ALL terms to match → too strict
      - plainto_tsquery treats terms as OR/proximity → much more forgiving
      - Handles stemming: "cars" matches "car", "Toyota" matches "toyota"
    """
    import psycopg2.extras
    from services.database_service import _get_connection

    q_lower = query.lower()

    # Tier 3: broad "show everything" queries
    if any(p in q_lower for p in ["show all", "list all", "all rows",
                                   "display all", "show me all", "get all",
                                   "show everything", "all data", "full data"]):
        return _get_all_rows(document_ids, top_k)

    # Extract meaningful terms
    search_terms = _extract_search_terms(query)
    if not search_terms:
        # No meaningful terms — return sample rows so user sees the data
        return _get_all_rows(document_ids, min(top_k, 10))

    # Tier 1: plainto_tsquery (phrase-based, lenient)
    phrase = " ".join(search_terms)

    try:
        results = _fts_search(phrase, document_ids, top_k)
        if results:
            logger.info("FTS search: '%s' → %d rows.", query[:60], len(results))
            return results
    except Exception as exc:
        logger.warning("FTS search failed: %s", exc)

    # Tier 2: ILIKE per term (substring match, most permissive)
    results = _ilike_search(search_terms, document_ids, top_k)
    if results:
        logger.info("ILIKE search: '%s' → %d rows.", query[:60], len(results))
        return results

    logger.info("No rows found for query: '%s'", query[:60])
    return []


def _fts_search(phrase: str, document_ids: list, top_k: int) -> list:
    """Full-text search using plainto_tsquery."""
    import psycopg2.extras
    from services.database_service import _get_connection

    scope = "AND e.document_id = ANY(%(doc_ids)s)" if document_ids else ""
    sql = f"""
        SELECT
            e.id, e.document_id, e.sheet_name, e.row_number,
            e.row_data, e.row_text, d.filename,
            ts_rank(to_tsvector('english', e.row_text),
                    plainto_tsquery('english', %(phrase)s)) AS rank
        FROM excel_rows e
        JOIN documents d ON d.id = e.document_id
        WHERE to_tsvector('english', e.row_text) @@ plainto_tsquery('english', %(phrase)s)
        {scope}
        ORDER BY rank DESC
        LIMIT %(top_k)s;
    """
    params = {"phrase": phrase, "top_k": top_k}
    if document_ids:
        params["doc_ids"] = document_ids

    with _get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    return _parse_rows(rows)


def _ilike_search(terms: list, document_ids: list, top_k: int) -> list:
    """Substring search using ILIKE — catches what FTS misses."""
    import psycopg2.extras
    from services.database_service import _get_connection

    # Build OR conditions for each term
    conditions = " OR ".join([f"e.row_text ILIKE %(t{i})s" for i in range(len(terms))])
    scope      = "AND e.document_id = ANY(%(doc_ids)s)" if document_ids else ""

    sql = f"""
        SELECT e.id, e.document_id, e.sheet_name, e.row_number,
               e.row_data, e.row_text, d.filename, 1.0 AS rank
        FROM excel_rows e
        JOIN documents d ON d.id = e.document_id
        WHERE ({conditions}) {scope}
        LIMIT %(top_k)s;
    """
    params = {f"t{i}": f"%{t}%" for i, t in enumerate(terms)}
    params["top_k"] = top_k
    if document_ids:
        params["doc_ids"] = document_ids

    with _get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    return _parse_rows(rows)


def _get_all_rows(document_ids: list, top_k: int) -> list:
    """Return all rows (or first top_k) — for broad queries."""
    import psycopg2.extras
    from services.database_service import _get_connection

    scope = "WHERE e.document_id = ANY(%s)" if document_ids else ""
    sql = f"""
        SELECT e.id, e.document_id, e.sheet_name, e.row_number,
               e.row_data, e.row_text, d.filename, 1.0 AS rank
        FROM excel_rows e
        JOIN documents d ON d.id = e.document_id
        {scope}
        ORDER BY e.document_id, e.sheet_name, e.row_number
        LIMIT %s;
    """
    params = ([document_ids, top_k] if document_ids else [top_k])
    if document_ids:
        params = (document_ids, top_k)
    else:
        params = (top_k,)

    with _get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = cur.fetchall()

    return _parse_rows(rows)


def _parse_rows(rows) -> list:
    """Convert DB rows to clean dicts."""
    results = []
    for r in rows:
        row = dict(r)
        if isinstance(row.get("row_data"), str):
            row["row_data"] = json.loads(row["row_data"])
        results.append(row)
    return results


def _fallback_search(query: str, document_ids: list, top_k: int) -> list:
    """
    Simple ILIKE fallback when full-text search fails
    (e.g. special characters in the query).
    """
    import psycopg2.extras
    from services.database_service import _get_connection

    terms = _extract_search_terms(query)
    if not terms:
        return []

    pattern = f"%{terms[0]}%"
    scope   = "AND e.document_id = ANY(%s)" if document_ids else ""
    params  = [pattern, document_ids, top_k] if document_ids else [pattern, top_k]

    sql = f"""
        SELECT e.id, e.document_id, e.sheet_name, e.row_number,
               e.row_data, e.row_text, d.filename, 1.0 AS rank
        FROM excel_rows e
        JOIN documents d ON d.id = e.document_id
        WHERE e.row_text ILIKE %s {scope}
        LIMIT %s;
    """

    with _get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            rows = [dict(r) for r in cur.fetchall()]

    for r in rows:
        if isinstance(r["row_data"], str):
            r["row_data"] = json.loads(r["row_data"])
    return rows


def _extract_search_terms(query: str) -> list:
    """
    Extract meaningful search terms from a query.
    Removes common filler words and returns tokens suitable for ts_query.
    """
    import re
    stop = {"is", "are", "the", "a", "an", "in", "of", "for", "to", "there",
            "any", "find", "show", "get", "what", "does", "do", "has", "have",
            "me", "this", "that", "which", "where", "how", "can", "could"}
    tokens = re.findall(r"[a-zA-Z0-9\-]+", query)
    return [t for t in tokens if t.lower() not in stop and len(t) >= 2]


# ---------------------------------------------------------------------------
# Aggregation with LLM-generated SQL (Type 2)
# ---------------------------------------------------------------------------

def answer_aggregation(query: str, document_id: int = None) -> dict:
    """
    Answer an aggregation question about Excel data using LLM-generated SQL.

    Flow:
        1. Fetch schema (sheet names, columns) for context
        2. LLM generates a SQL query for the excel_rows table
        3. Execute the SQL
        4. LLM generates a natural language summary

    Returns:
        dict with: answered, sql, result_rows, response
    """
    import psycopg2.extras
    from services.database_service import _get_connection, get_all_documents
    from services.llm_service import _call_ollama

    # Get schema context
    schema_parts = []
    docs = get_all_documents()
    for doc in docs:
        if document_id and doc["id"] != document_id:
            continue
        meta = {}
        try:
            from services.metadata_service import get_metadata
            meta = get_metadata(doc["id"])
        except Exception:
            pass
        if meta:
            schema_parts.append(f"File: {doc['filename']}")
            for k, v in meta.items():
                if "column" in k or "sheet" in k:
                    schema_parts.append(f"  {k}: {v}")

    schema_context = "\n".join(schema_parts) if schema_parts else "No schema available."

    # Prompt LLM to generate SQL
    sql_prompt = f"""You are a SQL expert. Generate a PostgreSQL SQL query to answer this question.

The data is stored in the excel_rows table with these columns:
  - document_id (INTEGER)
  - sheet_name  (TEXT)
  - row_number  (INTEGER)
  - row_data    (JSONB) — contains the actual column values as key-value pairs

Document schema context:
{schema_context}

To access a JSONB field, use: row_data->>'ColumnName'
To count rows: SELECT COUNT(*) FROM excel_rows WHERE ...
To sum: SELECT SUM((row_data->>'Amount')::numeric) FROM excel_rows WHERE ...

Question: {query}

Return ONLY the SQL query, no explanation, no markdown."""

    try:
        raw_sql = _call_ollama(sql_prompt).strip()
        # Clean up SQL (remove markdown code blocks if present)
        import re
        raw_sql = re.sub(r"```(?:sql)?", "", raw_sql).strip()

        # Safety: only allow SELECT queries
        if not raw_sql.upper().startswith("SELECT"):
            return {
                "answered": False,
                "response": "Could not generate a safe SQL query for this question.",
                "sql":      raw_sql,
                "result_rows": []
            }

        # Execute the SQL
        with _get_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(raw_sql)
                result_rows = [dict(r) for r in cur.fetchall()]

        # LLM generates natural language summary
        summary_prompt = f"""Given this question: "{query}"
And these SQL results: {json.dumps(result_rows[:10])}
Write a clear, concise one-sentence answer in plain English."""

        response = _call_ollama(summary_prompt).strip()

        return {
            "answered":    True,
            "sql":         raw_sql,
            "result_rows": result_rows,
            "response":    response,
        }

    except Exception as exc:
        logger.warning("Aggregation query failed: %s", exc)
        return {
            "answered": False,
            "response": f"Could not compute the answer: {exc}",
            "sql":      "",
            "result_rows": []
        }


# ---------------------------------------------------------------------------
# Query intent detection
# ---------------------------------------------------------------------------

LOOKUP_KEYWORDS = {
    "find", "search", "show", "get", "is", "are", "exists", "exist",
    "look", "check", "there", "any", "which", "locate", "list",
    "order", "invoice", "bill", "record", "entry", "row",
}

AGGREGATION_KEYWORDS = {
    "how many", "count", "total", "sum", "average", "max", "min",
    "highest", "lowest", "most", "least", "percentage", "percent",
    "calculate", "compute", "aggregate",
}


def detect_excel_intent(query: str) -> str:
    """
    Classify an Excel-related query into intent type.

    Returns:
        "lookup"      — row search
        "aggregation" — computed answer
        "metadata"    — structure question (handled by metadata_service)
        "unknown"     — cannot determine
    """
    q = query.lower()

    from services.metadata_service import is_metadata_question
    if is_metadata_question(query):
        return "metadata"

    if any(phrase in q for phrase in AGGREGATION_KEYWORDS):
        return "aggregation"

    if any(word in q.split() for word in LOOKUP_KEYWORDS):
        return "lookup"

    return "lookup"   # default to lookup for Excel queries